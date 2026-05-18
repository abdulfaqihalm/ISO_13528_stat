#!/usr/bin/env python3
"""
Algorithm A — ISO 13528-style robust mean and robust standard deviation.

Reproduces the behaviour of the Excel workbook Algorithm-A.xls, including the
VBA macro that sorts input values, performs iterative winsorization, and reports
the robust mean (AlgA-mean) and robust standard deviation (AlgA-std).

Usage
-----
    python alg_a.py input.csv --value-column measurement --output alg_a_output.csv

See --help for all options.
"""

from __future__ import annotations

import argparse
import csv
import math
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Core statistical helpers
# ---------------------------------------------------------------------------


def median_excel(values: list[float]) -> float:
    """Return the median, matching Excel MEDIAN behaviour.

    For an even-length list the average of the two middle values is returned,
    which is identical to Excel's MEDIAN and Python's statistics.median.
    """
    n = len(values)
    if n == 0:
        raise ValueError("Cannot compute median of an empty list.")
    s = sorted(values)
    mid = n // 2
    if n % 2 == 1:
        return float(s[mid])
    return (s[mid - 1] + s[mid]) / 2.0


def sample_std(values: list[float]) -> float:
    """Return the sample standard deviation (denominator n-1), matching Excel STDEV.

    Returns 0.0 when n < 2, matching the graceful fallback used in the
    iteration when the dataset degenerates to a single value.
    """
    n = len(values)
    if n < 2:
        return 0.0
    mean = sum(values) / n
    variance = sum((v - mean) ** 2 for v in values) / (n - 1)
    return math.sqrt(variance)


# ---------------------------------------------------------------------------
# Algorithm A
# ---------------------------------------------------------------------------


def algorithm_a(
    values: list[float],
    compat_excel: bool = True,
    max_iter: int = 1000,
) -> dict:
    """Run ISO 13528 Algorithm A (robust mean and standard deviation).

    Parameters
    ----------
    values:
        Numeric measurement values. Blanks/non-numerics must be removed before
        calling this function.
    compat_excel:
        When True (default) the convergence tolerance matches the workbook
        exactly: ``abs(x_new - x_old) < x_new / 1_000_000``.
        When False a safer form is used:
        ``abs(x_new - x_old) < max(abs(x_new) / 1_000_000, 1e-12)``.
    max_iter:
        Safety cap on the number of iterations. The workbook sample converges
        in 17 iterations; 1000 is a generous ceiling.

    Returns
    -------
    dict with keys:
        sorted_values   – input values sorted ascending (workbook column B)
        abs_dev         – |value_i - starting_median| (workbook column C)
        adjusted        – winsorized values from the final iteration (column D)
        median          – starting median, D1
        s_mad           – starting robust scale 1.483·MAD, D2
        n               – count of values, D3
        robust_mean     – AlgA-mean, F1
        robust_std      – AlgA-std, F2
        iterations      – number of iterations, I1
    """
    values = sorted(values)
    n = len(values)
    if n == 0:
        raise ValueError("No numeric measurement values found.")

    # Starting estimates (workbook cells D1 and D2)
    starting_median = median_excel(values)
    abs_dev = [abs(v - starting_median) for v in values]
    s_mad = 1.483 * median_excel(abs_dev)

    x: float = starting_median
    s: float = s_mad
    adjusted: list[float] = values[:]
    iterations: int = 0

    while True:
        d = 1.5 * s
        lo = x - d
        hi = x + d
        adjusted = [min(max(v, lo), hi) for v in values]

        x_old = x
        x = sum(adjusted) / n
        s = 1.134 * sample_std(adjusted) if n > 1 else 0.0
        iterations += 1

        if compat_excel:
            tolerance = x / 1_000_000
        else:
            tolerance = max(abs(x) / 1_000_000, 1e-12)

        if abs(x - x_old) < tolerance:
            break
        if iterations >= max_iter:
            raise RuntimeError(f"Algorithm A did not converge after {max_iter} iterations.")

    return {
        "sorted_values": values,
        "abs_dev": abs_dev,
        "adjusted": adjusted,
        "median": starting_median,
        "s_mad": s_mad,
        "n": n,
        "robust_mean": x,
        "robust_std": s,
        "iterations": iterations,
    }


# ---------------------------------------------------------------------------
# CSV / XLSX I/O helpers
# ---------------------------------------------------------------------------


def _fmt(value: float, precision: int) -> str:
    """Format a float to *precision* significant digits without trailing zeros."""
    if math.isnan(value):
        return ""
    return f"{value:.{precision}g}"


def load_csv(
    path: Path,
    value_column: str | int,
    zscore_column: str | int | None = None,
) -> tuple[list[float], list[str | None]]:
    """Read *value_column* (and optionally *zscore_column*) from a CSV file.

    Parameters
    ----------
    value_column:
        Column name (case-insensitive header match) or 1-based integer index.
    zscore_column:
        Optional column name or 1-based integer index for the z-score input
        values (equivalent to workbook column G).

    Returns
    -------
    (values, z_inputs)
        *values* contains only successfully parsed floats.
        *z_inputs* is a list of the same length as the CSV data rows;
        each entry is a string (raw value) or None if no zscore column was
        supplied.  Invalid z-score entries are preserved as the raw string so
        the caller can decide what to emit.
    """
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV file '{path}' is empty.")

    header = rows[0]
    data_rows = rows[1:]

    def resolve_col(spec: str | int) -> int:
        """Return 0-based column index from a name or 1-based integer."""
        if isinstance(spec, int):
            idx = spec - 1
            if idx < 0 or idx >= len(header):
                raise ValueError(
                    f"Column index {spec} is out of range (CSV has {len(header)} columns)."
                )
            return idx
        # Name lookup (case-insensitive)
        lower_header = [h.strip().lower() for h in header]
        name_lower = spec.strip().lower()
        if name_lower in lower_header:
            return lower_header.index(name_lower)
        # Try interpreting as a 1-based integer string
        try:
            idx = int(spec) - 1
            if 0 <= idx < len(header):
                return idx
        except (ValueError, TypeError):
            pass
        raise ValueError(f"Column '{spec}' not found in CSV header: {header}")

    val_idx = resolve_col(value_column)
    z_idx = resolve_col(zscore_column) if zscore_column is not None else None

    values: list[float] = []
    z_inputs: list[str | None] = []

    for row in data_rows:
        # Extend short rows with empty strings
        while len(row) <= max(val_idx, z_idx if z_idx is not None else 0):
            row.append("")

        raw = row[val_idx].strip()
        if raw == "":
            # Skip blank measurement values
            continue
        try:
            values.append(float(raw))
        except ValueError:
            # Skip non-numeric measurement values
            continue

        z_raw = row[z_idx].strip() if z_idx is not None else None
        z_inputs.append(z_raw)

    return values, z_inputs


def write_csv(
    path: Path,
    result: dict,
    z_inputs: list[str | None],
    precision: int = 15,
) -> None:
    """Write the Algorithm A result to *path* in the workbook-like CSV layout.

    Layout (columns A–I, 1-indexed):
    ┌────┬──────────┬──────────┬──────────┬───┬───────────┬────────────┬──────────┬────────────┐
    │ A  │    B     │    C     │    D     │ E │     F     │     G      │    H     │     I      │
    ├────┼──────────┼──────────┼──────────┼───┼───────────┼────────────┼──────────┼────────────┤
    │    │          │ median   │ <median> │   │ AlgA-mean │ <mean>     │iterations│ <iters>    │
    │    │          │ s(MAD)   │ <s_mad>  │   │ AlgA-std  │ <std>      │          │            │
    │    │          │ number   │ <n>      │   │           │            │          │            │
    │    │ data     │ abs.dev. │ number   │   │           │ z_input    │ Zscore   │            │
    │  1 │ <val1>   │ <dev1>   │ <adj1>   │   │           │ <z1>       │ <zs1>    │            │
    └────┴──────────┴──────────┴──────────┴───┴───────────┴────────────┴──────────┴────────────┘
    """
    robust_mean = result["robust_mean"]
    robust_std = result["robust_std"]
    has_z = any(z is not None for z in z_inputs)

    def fv(v: float) -> str:
        return _fmt(v, precision)

    def zscore(z_raw: str | None) -> str:
        if z_raw is None or z_raw == "":
            return ""
        try:
            z_val = float(z_raw)
        except ValueError:
            return z_raw  # preserve non-numeric as-is
        if robust_std == 0.0:
            return ""
        return fv((z_val - robust_mean) / robust_std)

    rows: list[list[str]] = []

    # Header row (column labels)
    rows.append(["A", "B", "C", "D", "E", "F", "G", "H", "I"])

    # Summary rows 1–3
    rows.append(
        [
            "",
            "",
            "median",
            fv(result["median"]),
            "AlgA-mean",
            fv(robust_mean),
            "",
            "iterations:",
            str(result["iterations"]),
        ]
    )
    rows.append(
        [
            "",
            "",
            "s(MAD)",
            fv(result["s_mad"]),
            "AlgA-std",
            fv(robust_std),
            "",
            "",
            "",
        ]
    )
    rows.append(
        [
            "",
            "",
            "number",
            str(result["n"]),
            "",
            "",
            "",
            "",
            "",
        ]
    )

    # Column header row
    rows.append(
        [
            "",
            "data",
            "abs. dev.",
            "number",
            "",
            "",
            "z_input" if has_z else "",
            "Zscore" if has_z else "",
            "",
        ]
    )

    # Data rows
    sorted_vals = result["sorted_values"]
    abs_devs = result["abs_dev"]
    adjusted = result["adjusted"]

    for i, (v, dev, adj) in enumerate(zip(sorted_vals, abs_devs, adjusted, strict=False), start=1):
        z_raw = z_inputs[i - 1] if i - 1 < len(z_inputs) else None
        rows.append(
            [
                str(i),
                fv(v),
                fv(dev),
                fv(adj),
                "",
                "",
                z_raw if z_raw is not None else "",
                zscore(z_raw) if has_z else "",
                "",
            ]
        )

    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)


def write_xlsx(path: Path, result: dict, z_inputs: list[str | None], precision: int = 15) -> None:
    """Write the Algorithm A result to an Excel workbook (.xlsx).

    Requires openpyxl. If not installed this function raises ImportError with
    a helpful message.
    """
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AlgA"

    robust_mean = result["robust_mean"]
    robust_std = result["robust_std"]
    has_z = any(z is not None for z in z_inputs)

    def fv(v: float) -> float:
        return round(v, precision)

    def zscore(z_raw: str | None) -> float | str:
        if z_raw is None or z_raw == "":
            return ""
        try:
            z_val = float(z_raw)
        except ValueError:
            return z_raw
        if robust_std == 0.0:
            return ""
        return (z_val - robust_mean) / robust_std

    bold = Font(bold=True)

    # Row 1 summary
    ws["C1"] = "median"
    ws["D1"] = fv(result["median"])
    ws["E1"] = "AlgA-mean"
    ws["F1"] = fv(robust_mean)
    ws["H1"] = "iterations:"
    ws["I1"] = result["iterations"]
    ws["C1"].font = bold
    ws["E1"].font = bold

    # Row 2
    ws["C2"] = "s(MAD)"
    ws["D2"] = fv(result["s_mad"])
    ws["E2"] = "AlgA-std"
    ws["F2"] = fv(robust_std)
    ws["C2"].font = bold
    ws["E2"].font = bold

    # Row 3
    ws["C3"] = "number"
    ws["D3"] = result["n"]
    ws["C3"].font = bold

    # Row 4 column headers
    for col, label in [("B", "data"), ("C", "abs. dev."), ("D", "number")]:
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = bold
    if has_z:
        ws["G4"] = "z_input"
        ws["H4"] = "Zscore"
        ws["G4"].font = bold
        ws["H4"].font = bold

    # Data rows starting at row 5
    sorted_vals = result["sorted_values"]
    abs_devs = result["abs_dev"]
    adjusted = result["adjusted"]

    for i, (v, dev, adj) in enumerate(zip(sorted_vals, abs_devs, adjusted, strict=False), start=1):
        row = i + 4
        z_raw = z_inputs[i - 1] if i - 1 < len(z_inputs) else None
        ws.cell(row=row, column=1).value = i  # A: row number
        ws.cell(row=row, column=2).value = fv(v)  # B: sorted value
        ws.cell(row=row, column=3).value = fv(dev)  # C: abs deviation
        ws.cell(row=row, column=4).value = fv(adj)  # D: adjusted value
        if has_z and z_raw is not None:
            ws.cell(row=row, column=7).value = z_raw  # G: z_input
            ws.cell(row=row, column=8).value = zscore(z_raw)  # H: Zscore

    wb.save(path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="alg_a.py",
        description=(
            "Algorithm A — ISO 13528-style robust mean and standard deviation.\n"
            "Reproduces the behaviour of the Excel workbook Algorithm-A.xls."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "input",
        type=Path,
        help="Path to the input CSV file.",
    )
    p.add_argument(
        "--value-column",
        default="1",
        metavar="COL",
        help=(
            "Name (case-insensitive) or 1-based integer index of the column "
            "containing measurement values. Default: 1 (first column)."
        ),
    )
    p.add_argument(
        "--zscore-column",
        default=None,
        metavar="COL",
        help=(
            "Optional column name or 1-based integer index of a second column "
            "whose values are used to compute z-scores (workbook column G). "
            "If omitted, no z-scores are written."
        ),
    )
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        metavar="FILE",
        help=(
            "Path for the CSV output file. "
            "Default: <input_stem>_alg_a_output.csv in the same directory."
        ),
    )
    p.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        metavar="FILE",
        help="Optional path to also write an XLSX workbook (requires openpyxl).",
    )
    p.add_argument(
        "--compat-excel",
        action="store_true",
        default=True,
        help=(
            "Use the exact workbook convergence rule "
            "abs(x_new - x_old) < x_new / 1_000_000. "
            "This is the default. Use --no-compat-excel for a safer variant."
        ),
    )
    p.add_argument(
        "--no-compat-excel",
        dest="compat_excel",
        action="store_false",
        help=(
            "Use the safer convergence rule "
            "abs(x_new - x_old) < max(abs(x_new) / 1_000_000, 1e-12)."
        ),
    )
    p.add_argument(
        "--max-iter",
        type=int,
        default=1000,
        metavar="N",
        help="Maximum number of iterations before raising an error. Default: 1000.",
    )
    p.add_argument(
        "--precision",
        type=int,
        default=15,
        metavar="DIGITS",
        help="Number of significant digits in output. Default: 15.",
    )
    return p


def _resolve_value_column(spec: str) -> str | int:
    """Return an int if *spec* is a pure integer string, otherwise return as-is."""
    try:
        return int(spec)
    except (ValueError, TypeError):
        return spec


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    input_path: Path = args.input
    if not input_path.exists():
        print(f"Error: input file '{input_path}' not found.", file=sys.stderr)
        return 1

    # Resolve output path
    output_path: Path = args.output or input_path.with_name(input_path.stem + "_alg_a_output.csv")

    # Parse column specs
    value_col = _resolve_value_column(args.value_column)
    zscore_col = _resolve_value_column(args.zscore_column) if args.zscore_column else None

    # Load data
    try:
        values, z_inputs = load_csv(input_path, value_col, zscore_col)
    except (ValueError, OSError) as exc:
        print(f"Error reading CSV: {exc}", file=sys.stderr)
        return 1

    if not values:
        print("Error: No numeric measurement values found in the input file.", file=sys.stderr)
        return 1

    # Run Algorithm A
    try:
        result = algorithm_a(values, compat_excel=args.compat_excel, max_iter=args.max_iter)
    except (ValueError, RuntimeError) as exc:
        print(f"Error running Algorithm A: {exc}", file=sys.stderr)
        return 1

    # Write CSV output
    try:
        write_csv(output_path, result, z_inputs, precision=args.precision)
    except OSError as exc:
        print(f"Error writing output CSV: {exc}", file=sys.stderr)
        return 1

    # Write XLSX output (optional)
    if args.output_xlsx:
        try:
            write_xlsx(args.output_xlsx, result, z_inputs, precision=args.precision)
        except (ImportError, OSError) as exc:
            print(f"Warning: could not write XLSX: {exc}", file=sys.stderr)

    # Print summary to stdout
    print(f"Algorithm A complete — {result['iterations']} iteration(s).")
    print(f"  n           = {result['n']}")
    print(f"  median      = {result['median']:.{args.precision}g}")
    print(f"  s(MAD)      = {result['s_mad']:.{args.precision}g}")
    print(f"  AlgA-mean   = {result['robust_mean']:.{args.precision}g}")
    print(f"  AlgA-std    = {result['robust_std']:.{args.precision}g}")
    print(f"  Output CSV  → {output_path}")
    if args.output_xlsx:
        print(f"  Output XLSX → {args.output_xlsx}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
